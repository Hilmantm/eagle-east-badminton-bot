from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
import pandas as pd

SESSION_1 = 1
SESSION_2 = 2
SESSION_3 = 3
SESSION_4 = 4
SESSION_5 = 5

day_index = {
    "Senin": 0,
    "Selasa": 1,
    "Rabu": 2,
    "Kamis": 3,
    "Jum'at": 4,
    "Sabtu": 5,
    "Minggu": 6
}

current_schedule = {
    "Senin": ["", "", "", "", ""],
    "Selasa": ["", "", "", "", ""],
    "Rabu": ["", "", "", "", ""],
    "Kamis": ["", "", "", "", ""],
    "Jum'at": ["", "", "", "", ""],
    "Sabtu": ["", "", "", "", ""],
    "Minggu": ["", "", "", "", ""]
}

def register_schedule_request(message):
    request = message.text.split("_")
    if len(request) == 4 and request[0].lower() == "daftar":
        return True
    return False


def schedule_workbook():
    wb = Workbook()
    ws = wb.active

    df = df_schedule_template()
    fill_current_schedule(df)

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    workbook_schedule_format(ws)

    wb.save("current-schedule.xlsx")

    return wb


def fill_current_schedule(df):
    for schedule in current_schedule:
        for index, name in enumerate(current_schedule[schedule]):
            if index != 0:
                df.loc[day_index[schedule]][index] = current_schedule[schedule][index]

    return df


'''
If want to add name to specific column
df.loc[0][1] = "Pak Parman" => senin sesi 1
df.loc[5][5] = "Pak Asep"
First array represented row in df
Second array represented session or column in df 
'''
def df_schedule_template():
    df = pd.DataFrame(
        columns=["Jadwal/Hari", "07.00 - 10.00", "10.00 - 13.00", "13.00 - 16.00", "16.00 - 19.00", "19.00 - 22.00"])
    df["Jadwal/Hari"] = ["Senin", "Selasa", "Rabu", "Kamis", "Jum'at", "Sabtu", "Minggu"]
    return df

def workbook_schedule_format(ws):
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18

    # make session bold
    for cell in ws['1:1']:
        cell.font = Font(bold=True)

    # make day bold
    for row in ws['A']:
        row.font = Font(bold=True)

    # make text center
    cols = ['B', 'C', 'D', 'E', 'F']
    for col in cols:
        for row in ws[col]:
            row.alignment = Alignment(horizontal="center")